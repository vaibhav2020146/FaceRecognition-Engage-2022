from email.mime import application
from operator import truediv
from flask import Flask,render_template,redirect,request
import openpyxl
from werkzeug.utils import secure_filename
from werkzeug.datastructures import  FileStorage
import cv2
import numpy as np
import face_recognition
import os
from os.path import exists
from datetime import datetime,date
import pandas as pd
from openpyxl import load_workbook
import csv
import dlib
import geocoder
latitude = 0
longitude = 0

application=Flask(__name__)
def findEncodings(images):
    #will do encoding of the images and store it in the encodeListKnown
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList
    
def markAttendance_for_entry(name):
    #will mark the attendance of the student and save it in the csv file
    #df = pd.read_excel("Attendance.xlsx", usecols = ['Username'])
    with open('Attendance.csv', 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = datetime.now()
            date_today=date.today();
            dtString = now.strftime('%H:%M:%S')
            convert_to_format=date_today.strftime("%d-%m-%Y")
            wb=openpyxl.load_workbook('Attendance.xlsx')
            sh1=wb['Sheet1']
            row=sh1.max_row
            column=sh1.max_column
            is_added_entry=False
            for i in range(1,row+1):
                if (sh1.cell(row=i,column=1).value==name) and (sh1.cell(row=i,column=4).value==convert_to_format):
                    is_added_entry=True

            if(is_added_entry==False):
                sh1.cell(row=row+1,column=1,value=name)
                sh1.cell(row=row+1,column=2,value=dtString)
                sh1.cell(row=row+1,column=3,value='NA')
                sh1.cell(row=row+1,column=4,value=convert_to_format)
            wb.save('Attendance.xlsx')
            return (True,"Arival Attendance marked for "+name)
    return (False,"")

def markAttendance_for_leaving(name):
    check_if_username_has_entered=False
    wb=openpyxl.load_workbook('Attendance.xlsx')
    sh1=wb['Sheet1']
    row=sh1.max_row
    for i in range(1,row+1):
        if (sh1.cell(row=i,column=1).value==name):
            check_if_username_has_entered=True
            break
    #wb.save('Attendance.xlsx')
    if(check_if_username_has_entered==False):
        return (True,"You have not entered the attendance")
    with open('Attendance.csv', 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = datetime.now()
            date_today=date.today();
            dtString = now.strftime('%H:%M:%S')
            convert_to_format=date_today.strftime("%d-%m-%Y")
            wb=openpyxl.load_workbook('Attendance.xlsx')
            sh1=wb['Sheet1']
            row=sh1.max_row
            is_added_leave=False
            for i in range(1,row+1):
                if (sh1.cell(row=i,column=1).value==name) and (sh1.cell(row=i,column=4).value==convert_to_format and sh1.cell(row=i,column=3).value=='NA'):
                    is_added_leave=True

            if(is_added_leave==False):
                #sh1.cell(row=row+1,column=3,value=dtString)
                row_num=-1
                for i in range(1,row+1):
                    if (sh1.cell(row=i,column=1).value==name):
                        row_num=i
                if(row_num!=-1):
                    position="C"+str(row_num)
                    sh1[position].value=dtString
            wb.save('Attendance.xlsx')
            return (True,"Departure Attendance marked for "+name)
    return (False,"")


@application.route('/clear')
def clear():
    wb=openpyxl.load_workbook('Attendance.xlsx')
    sh1=wb['Sheet1']
    row=sh1.max_row
    for i in range(2,row+1):
        for cell in sh1[i]:
            cell.value=''
    sh1.cell(row=1,column=1,value='Name')
    sh1.cell(row=1,column=2,value='Entry Time')
    sh1.cell(row=1,column=3,value='Leave Time')
    sh1.cell(row=1,column=4,value='Date')
    wb.save('Attendance.xlsx')
    return render_template('/attendance_record.html')

@application.route('/table')
def table():
    df = pd.read_excel('Attendance.xlsx')
    df.to_excel('Attendance.xlsx', index=None)
    data = pd.read_excel('Attendance.xlsx')
    return render_template('/attendance_record.html', tables=[data.to_html()], titles=[''])



APP_ROOT = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLD = 'static/uploads'
UPLOAD_FOLDER = os.path.join(APP_ROOT, UPLOAD_FOLD)
application.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@application.route('/location', methods=['GET', 'POST'])
def location():
    if request.method == 'POST':
        global longitude,latitude
        longitude=request.form['longitude']
        latitude=request.form['latitude']
        return render_template('/admin_page.html',info="Location Updated")
    return render_template('/admin_page.html')

@application.route('/confirm_location', methods=['GET', 'POST'])
def confirm_location():
    g = geocoder.ip('me')
    if request.method == 'POST':
        longitude_cor = request.form['longitude_cord']
        latitude_cor = request.form['latitude_cord']
        if(float(latitude_cor)-5.0<=float(g.latlng[0])<=float(latitude_cor)+5.0 and float(longitude_cor)-5.0<=float(g.latlng[1])<=float(longitude_cor)+5.0):
            if(((float(longitude)-5.0)<=float(longitude_cor)<=(float(longitude)+5.0)) and ((float(latitude)-5.0)<=float(latitude_cor)<=(float(latitude)+5.0))):
                return render_template('/mark_attendance.html')
            else:
                return render_template('/check_location_of_user.html',info="Location is not correct")
        else:
            return render_template('/check_location_of_user.html',info="You Are Not Present At The Current Location")   
    return render_template('/check_location_of_user.html')


@application.route('/form_signup',methods=['GET','POST'])
def form_signup():
    if request.method=='POST':
        uname=request.form['username']
        passw=request.form['password']
        mail=request.form['mail']
        wb=openpyxl.load_workbook('database.xlsx')
        sh1=wb['Sheet1']
        row=sh1.max_row
        check_if_username_exists=False
        for i in range(1,row+1):
            if uname == sh1.cell(row=i,column=1).value:
                check_if_username_exists=True
                break
                #return render_template('/login_page.html',info='Invalid Username')
        if(check_if_username_exists==True):
            wb.save('database.xlsx')
            return render_template('/signup_page.html',info='Username Already Exists')
        else:
            #database[uname]=[passw,mail]
            sh1.cell(row=row+1,column=1,value=uname)
            sh1.cell(row=row+1,column=2,value=passw)
            sh1.cell(row=row+1,column=3,value=mail)
            wb.save('database.xlsx')
            return render_template('/signup_page.html',info='Signup Successful')
    return render_template('/signup_page.html')


@application.route('/form_login',methods=['GET','POST'])
def form_login():
    #for logging in the the system
    if request.method=='POST':
        name1=request.form['username']
        pwd=request.form['password']
        #to confirm if admin is accessing the system
        if(name1=='admin' and pwd=='ms-engage'):
            return render_template('/admin_page.html')
        elif (name1=='admin' and pwd!='ms-engage'):
            return render_template('/login_page.html',info='Invalid Password')
        wb=openpyxl.load_workbook('database.xlsx')
        sh1=wb['Sheet1']
        row=sh1.max_row
        check_if_username_exists=False
        for i in range(1,row+1):
            if name1 == sh1.cell(row=i,column=1).value:
                check_if_username_exists=True
                break
                #return render_template('/login_page.html',info='Invalid Username')
        if(check_if_username_exists==False):
            wb.save('database.xlsx')
            return render_template('/login_page.html',info='Invalid Username')
        for i in range(1,row+1):
            if (name1 == sh1.cell(row=i,column=1).value) and (pwd == str(sh1.cell(row=i,column=2).value)):
                wb.save('database.xlsx')
                return render_template('/check_location_of_user.html')
            elif (name1 == sh1.cell(row=i,column=1).value and pwd != str(sh1.cell(row=i,column=2).value)):
                wb.save('database.xlsx')
                return render_template('/login_page.html',info='Invalid Password')
        wb.save('database.xlsx')
    return render_template('/login_page.html')


@application.route('/uploader', methods = ['GET', 'POST'])
def upload_file():
    #used to upload the image and save it in the same folder in which our project is
   if request.method == 'POST':
      f = request.files['file']
      #f.save(secure_filename(f.filename))
      f.save(os.path.join(application.config['UPLOAD_FOLDER'], secure_filename(f.filename)))
      myList = os.listdir(path)
      for cl in myList:
          curImg = cv2.imread(f'{path}/{cl}')
          images.append(curImg)
          classNames.append(os.path.splitext(cl)[0])
      #keep a close check here.
      global encodeListKnown
      encodeListKnown = findEncodings(images)
      return render_template('/login_page.html',info='Photo Uploaded')

@application.route("/")
def home():
    #open the home page of the website
    #change it to index.html test.html is for testing only
    return render_template("index.html")

@application.route("/about", methods=['GET', 'POST'])
def about():
    return render_template("/about.html") 

@application.route("/leaving_out", methods=['GET', 'POST'])
def leaving_out():
    human_is_real=True
    if human_is_real==True:
        cap = cv2.VideoCapture(0)
        while True:
            success, img = cap.read()
            # img = captureScreen()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
            facesCurFrame = face_recognition.face_locations(imgS)
            encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
        
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            to_close_the_web_cam=(False,"")
            for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
                matchIndex = np.argmin(faceDis)
            
                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    to_close_the_web_cam=markAttendance_for_leaving(name)
                    if(to_close_the_web_cam[0]):
                        break
            cv2.imshow('Webcam', img)
            cv2.waitKey(1)
            if(to_close_the_web_cam[0]):
                break
        cap.release()
        cv2.destroyAllWindows()
        return render_template("/mark_attendance.html",info=to_close_the_web_cam[1])


@application.route("/login")
def login():
    #open the camera and takes the attendence
    #if error occur because of it then put it as same as in FaceDetect.py code
    human_is_real=True
    if human_is_real==True:
        cap = cv2.VideoCapture(0)
        while True:
            success, img = cap.read()
            # img = captureScreen()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
            facesCurFrame = face_recognition.face_locations(imgS)
            encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
        
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            to_close_the_web_cam=(False,"")
            for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
                matchIndex = np.argmin(faceDis)
            
                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    to_close_the_web_cam=markAttendance_for_entry(name)
                    if(to_close_the_web_cam[0]):
                        break
            cv2.imshow('Webcam', img)
            cv2.waitKey(1)
            if(to_close_the_web_cam[0]):
                break
        cap.release()
        cv2.destroyAllWindows()
        return render_template("/mark_attendance.html",info=to_close_the_web_cam[1])


path = "static/uploads"
images = []
classNames = []
myList = os.listdir(path)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
encodeListKnown = findEncodings(images)

if __name__=="__main__":
    application.run(debug=False)
