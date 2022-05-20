from operator import truediv
from flask import Flask,render_template,redirect,request
import openpyxl
from werkzeug.utils import secure_filename
from werkzeug.datastructures import  FileStorage
import cv2
import numpy as np
import face_recognition
import os
from os.path import exists
from datetime import datetime,date
import pandas as pd
from openpyxl import load_workbook
import csv
import dlib
import math
BLINK_RATIO_THRESHOLD = 4.8

app=Flask(__name__)
database={'vaibhav':['1234','vaibhavrajpal26@gmail.com'],'vrinda':['5678','ab@yahoo.com'],'shivam':['abcd','xyz@gmail.com'],'elon':['tesla','spacex@gmail.com']}
#will make database like key would be the username and value would be list of password and mail

def midpoint(point1 ,point2):
    return (point1.x + point2.x)/2,(point1.y + point2.y)/2

def euclidean_distance(point1 , point2):
    return math.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

def get_blink_ratio(eye_points, facial_landmarks):
    
    #loading all the required points
    corner_left  = (facial_landmarks.part(eye_points[0]).x, 
                    facial_landmarks.part(eye_points[0]).y)
    corner_right = (facial_landmarks.part(eye_points[3]).x, 
                    facial_landmarks.part(eye_points[3]).y)
    
    center_top    = midpoint(facial_landmarks.part(eye_points[1]), 
                             facial_landmarks.part(eye_points[2]))
    center_bottom = midpoint(facial_landmarks.part(eye_points[5]), 
                             facial_landmarks.part(eye_points[4]))

    #calculating distance
    horizontal_length = euclidean_distance(corner_left,corner_right)
    vertical_length = euclidean_distance(center_top,center_bottom)

    ratio = horizontal_length / vertical_length

    return ratio

def check_if_human_is_real():
    cap = cv2.VideoCapture(0)
    cv2.namedWindow('BlinkDetector')
    #-----Step 3: Face detection with dlib----
    detector = dlib.get_frontal_face_detector()
    #-----Step 4: Detecting Eyes using landmarks in dlib-----
    predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")
    #these landmarks are based on the image above 
    left_eye_landmarks  = [36, 37, 38, 39, 40, 41]
    right_eye_landmarks = [42, 43, 44, 45, 46, 47]
    while True:
        #capturing frame
        retval, frame = cap.read()

        #exit the application if frame not found
        if not retval:
            print("Can't receive frame (stream end?). Exiting ...")
            return False 

        #-----Step 2: converting image to grayscale-----
        #frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        #-----Step 3: Face detection with dlib-----
        #detecting faces in the frame 
        faces,_,_ = detector.run(image = frame, upsample_num_times = 0, adjust_threshold = 0.0)

        #-----Step 4: Detecting Eyes using landmarks in dlib-----
        for face in faces:
            landmarks = predictor(frame, face)
            #-----Step 5: Calculating blink ratio for one eye-----
            left_eye_ratio  = get_blink_ratio(left_eye_landmarks, landmarks)
            right_eye_ratio = get_blink_ratio(right_eye_landmarks, landmarks)
            blink_ratio     = (left_eye_ratio + right_eye_ratio) / 2

            if blink_ratio > BLINK_RATIO_THRESHOLD:
                #Blink detected! Do Something!
                cv2.putText(frame,"BLINKING",(10,50), cv2.FONT_HERSHEY_SIMPLEX,
                2,(255,255,255),2,cv2.LINE_AA)
                return True
        cv2.imshow('BlinkDetector', frame)
        key = cv2.waitKey(1)
        if key == 27:
            break
    cap.release()
    cv2.destroyAllWindows()
    return False


def findEncodings(images):
    #will do encoding of the images and store it in the encodeListKnown
    encodeList = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList
    
def markAttendance_for_entry(name):
    #will mark the attendance of the student and save it in the csv file
    #df = pd.read_excel("Attendance.xlsx", usecols = ['Username'])
    #print(df)
    with open('Attendance.csv', 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        print(nameList)
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = datetime.now()
            date_today=date.today();
            dtString = now.strftime('%H:%M:%S')
            convert_to_format=date_today.strftime("%d-%m-%Y")
            #f.writelines(f'{name},{dtString},{convert_to_format}')
            '''writer = pd.ExcelWriter('Attendance.xlsx', engine='openpyxl') 
            wb  = writer.book
            df = pd.DataFrame({'Username': [name],
            'Date': [convert_to_format],
            'Time': [dtString]})
            df.to_excel(writer, index=False)
            wb.save('Attendance.xlsx')'''
            wb=openpyxl.load_workbook('Attendance.xlsx')
            sh1=wb['Sheet1']
            row=sh1.max_row
            column=sh1.max_column
               
            '''is_added_leave=False
            for i in range(1,row+1):
                if (sh1.cell(row=i,column=1).value==name) and (sh1.cell(row=i,column=4).value==convert_to_format) and (sh1.cell(row=i,column=3).value=='NA'):
                    is_added_leave=True

            if(is_added_leave==False):
                #sh1.cell(row=row+1,column=3,value=dtString)
                row_num=-1
                for i in range(1,row+1):
                    if (sh1.cell(row=i,column=1).value==name):
                        row_num=i
                if(row_num!=-1):
                    position="C"+str(row_num)
                    sh1[position].value=dtString'''

            is_added_entry=False
            for i in range(1,row+1):
                if (sh1.cell(row=i,column=1).value==name) and (sh1.cell(row=i,column=4).value==convert_to_format):
                    is_added_entry=True

            if(is_added_entry==False):
                sh1.cell(row=row+1,column=1,value=name)
                sh1.cell(row=row+1,column=2,value=dtString)
                sh1.cell(row=row+1,column=3,value='NA')
                sh1.cell(row=row+1,column=4,value=convert_to_format)
            wb.save('Attendance.xlsx')
            return (True,"Attendance marked for "+name)
    return (False,"")

def markAttendance_for_leaving(name):
    #will mark the attendance of the student and save it in the csv file
    #df = pd.read_excel("Attendance.xlsx", usecols = ['Username'])
    #print(df)
    check_if_username_has_entered=False
    wb=openpyxl.load_workbook('Attendance.xlsx')
    sh1=wb['Sheet1']
    row=sh1.max_row
    for i in range(1,row+1):
        if (sh1.cell(row=i,column=1).value==name):
            check_if_username_has_entered=True
            break
    wb.save('Attendance.xlsx')
    if(check_if_username_has_entered==False):
        return (True,"You have not entered the attendance")
    with open('Attendance.csv', 'r+') as f:
        myDataList = f.readlines()
        nameList = []
        print(nameList)
        for line in myDataList:
            entry = line.split(',')
            nameList.append(entry[0])
        if name not in nameList:
            now = datetime.now()
            date_today=date.today();
            dtString = now.strftime('%H:%M:%S')
            convert_to_format=date_today.strftime("%d-%m-%Y")
            #f.writelines(f'{name},{dtString},{convert_to_format}')
            '''writer = pd.ExcelWriter('Attendance.xlsx', engine='openpyxl') 
            wb  = writer.book
            df = pd.DataFrame({'Username': [name],
            'Date': [convert_to_format],
            'Time': [dtString]})
            df.to_excel(writer, index=False)
            wb.save('Attendance.xlsx')'''
            wb=openpyxl.load_workbook('Attendance.xlsx')
            sh1=wb['Sheet1']
            row=sh1.max_row
            column=sh1.max_column
               
            is_added_leave=False
            for i in range(1,row+1):
                if (sh1.cell(row=i,column=1).value==name) and (sh1.cell(row=i,column=4).value==convert_to_format) and (sh1.cell(row=i,column=3).value=='NA'):
                    is_added_leave=True

            if(is_added_leave==False):
                #sh1.cell(row=row+1,column=3,value=dtString)
                row_num=-1
                for i in range(1,row+1):
                    if (sh1.cell(row=i,column=1).value==name):
                        row_num=i
                if(row_num!=-1):
                    position="C"+str(row_num)
                    sh1[position].value=dtString
            wb.save('Attendance.xlsx')
            return (True,"Attendance marked for "+name)
    return (False,"")


@app.route('/table')
def table():
    df = pd.read_excel('Attendance.xlsx')
    df.to_excel('Attendance.xlsx', index=None)
    data = pd.read_excel('Attendance.xlsx')
    return render_template('/attendance_record.html', tables=[data.to_html()], titles=[''])



APP_ROOT = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLD = 'static/uploads'
UPLOAD_FOLDER = os.path.join(APP_ROOT, UPLOAD_FOLD)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/form_signup',methods=['GET','POST'])
def form_signup():
    if request.method=='POST':
        uname=request.form['username']
        passw=request.form['password']
        mail=request.form['mail']
        wb=openpyxl.load_workbook('database.xlsx')
        sh1=wb['Sheet1']
        row=sh1.max_row
        check_if_username_exists=False
        for i in range(1,row+1):
            if uname == sh1.cell(row=i,column=1).value:
                check_if_username_exists=True
                break
                #return render_template('/login_page.html',info='Invalid Username')
        if(check_if_username_exists==True):
            wb.save('database.xlsx')
            return render_template('/login_page.html',info='Username Already Exists')
        else:
            #database[uname]=[passw,mail]
            sh1.cell(row=row+1,column=1,value=uname)
            sh1.cell(row=row+1,column=2,value=passw)
            sh1.cell(row=row+1,column=3,value=mail)
            wb.save('database.xlsx')
            path = "static/uploads"
            images = []
            classNames = []
            myList = os.listdir(path)
            print(myList)
            for cl in myList:
                curImg = cv2.imread(f'{path}/{cl}')
                images.append(curImg)
                classNames.append(os.path.splitext(cl)[0])
            print(classNames)
            encodeListKnown = findEncodings(images)
            print('Encoding Complete')
            return render_template('/signup_page.html',info='UserName Is Available. Now Upload Your Photo')
    return render_template('/signup_page.html')


@app.route('/form_login',methods=['GET','POST'])
def form_login():
    #for logging in the the system
    if request.method=='POST':
        name1=request.form['username']
        pwd=request.form['password']
        #name1='vaibhav'
        # #pwd='12345'
        print(name1,pwd)
        wb=openpyxl.load_workbook('database.xlsx')
        sh1=wb['Sheet1']
        row=sh1.max_row
        check_if_username_exists=False
        for i in range(1,row+1):
            if name1 == sh1.cell(row=i,column=1).value:
                check_if_username_exists=True
                break
                #return render_template('/login_page.html',info='Invalid Username')
        if(check_if_username_exists==False):
            wb.save('database.xlsx')
            return render_template('/login_page.html',info='Invalid Username')
        '''else:
            if database[name1][0]!=pwd:
                # after login gets successful it redirects to the page where attendence woulde be taken
                # # make the attendence webpage and change the below location to it
                return render_template('/login_page.html',info='Invalid Password')
            else:
                return render_template('/mark_attendance.html')'''
        for i in range(1,row+1):
            print(sh1.cell(row=i,column=1).value,sh1.cell(row=i,column=2).value)
            if (name1 == sh1.cell(row=i,column=1).value) and (pwd == str(sh1.cell(row=i,column=2).value)):
                wb.save('database.xlsx')
                return render_template('/mark_attendance.html')
            elif (name1 == sh1.cell(row=i,column=1).value and pwd != str(sh1.cell(row=i,column=2).value)):
                wb.save('database.xlsx')
                return render_template('/login_page.html',info='Invalid Password')
        wb.save('database.xlsx')
    return render_template('/login_page.html')


@app.route('/uploader', methods = ['GET', 'POST'])
def upload_file():
    #used to upload the image and save it in the same folder in which our project is
   if request.method == 'POST':
      f = request.files['file']
      #f.save(secure_filename(f.filename))
      f.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(f.filename)))
      #print(database)
      return render_template('/login_page.html',info='Signup Successful')

@app.route("/")
def home():
    #open the home page of the website
    return render_template("index.html")

@app.route("/about", methods=['GET', 'POST'])
def about():
    return render_template("/about.html") 

@app.route("/leaving_out", methods=['GET', 'POST'])
def leaving_out():
    if check_if_human_is_real()==True:
        cap = cv2.VideoCapture(0)
        while True:
            success, img = cap.read()
            # img = captureScreen()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
            facesCurFrame = face_recognition.face_locations(imgS)
            encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
        
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            to_close_the_web_cam=(False,"")
            for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
                # print(faceDis)
                matchIndex = np.argmin(faceDis)
            
                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    print(name)
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    to_close_the_web_cam=markAttendance_for_leaving(name)
                    if(to_close_the_web_cam[0]):
                        break
            cv2.imshow('Webcam', img)
            cv2.waitKey(1)
            if(to_close_the_web_cam[0]):
                break
        cap.release()
        cv2.destroyAllWindows()
        #return render_template("index.html")
        print(to_close_the_web_cam[1])
        return render_template("/mark_attendance.html",info=to_close_the_web_cam[1])


@app.route("/login")
def login():
    #open the camera and takes the attendence
    #if error occur because of it then put it as same as in FaceDetect.py code
    if check_if_human_is_real()==True:
        cap = cv2.VideoCapture(0)
        while True:
            success, img = cap.read()
            # img = captureScreen()
            imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
            facesCurFrame = face_recognition.face_locations(imgS)
            encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)
        
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            to_close_the_web_cam=(False,"")
            for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
                matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
                # print(faceDis)
                matchIndex = np.argmin(faceDis)
            
                if matches[matchIndex]:
                    name = classNames[matchIndex].upper()
                    print(name)
                    y1, x2, y2, x1 = faceLoc
                    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                    cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                    cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                    to_close_the_web_cam=markAttendance_for_entry(name)
                    if(to_close_the_web_cam[0]):
                        break
            cv2.imshow('Webcam', img)
            cv2.waitKey(1)
            if(to_close_the_web_cam[0]):
                break
        cap.release()
        cv2.destroyAllWindows()
        #return render_template("index.html")
        print(to_close_the_web_cam[1])
        return render_template("/mark_attendance.html",info=to_close_the_web_cam[1])


path = "static/uploads"
images = []
classNames = []
myList = os.listdir(path)
print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)
encodeListKnown = findEncodings(images)
print('Encoding Complete')

if __name__=="__main__":
    #database.clear()
    app.run(debug=True)
